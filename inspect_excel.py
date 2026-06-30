import openpyxl, sys, json

path = sys.argv[1] if len(sys.argv) > 1 else '新产品作战地图.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print('sheets:', wb.sheetnames)
all_text = []
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n--- Sheet: {sheet_name} ---')
    print(f'rows: {ws.max_row}, cols: {ws.max_column}')
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            if cell is not None:
                all_text.append(str(cell))

open('excel_text.txt','w',encoding='utf-8').write('\n'.join(all_text))
print('total text cells', len(all_text))
